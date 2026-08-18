# -*- coding: utf-8 -*-
"""
converter_os.py — Geoportal de Inspeção Rodoviária (RTA-MSI / Tocantins)

Lê a(s) planilha(s) de Controle de O.S.P. (fichas/OS/*.xlsm — uma aba
"REGIÃO XX - RESUMO" por região, uma linha por O.S.P.+trecho) e gera, por
região GEOGRÁFICA (a mesma numeração da ficha de inspeção: R1, R2, R3, R11,
R12, R13), um GeoJSON com uma feature por linha do shapefile de trechos
(camadas/R<região>_TRECHOS.shp, campo `Id`) que bater com o "N° TRECHO" da
O.S.P. — a geometria é o trecho INTEIRO (todos os S.R.E. daquele Id), sem
corte por km: diferente da ficha, a O.S.P. não referencia sub-trecho.

Região de manutenção x restauração — mesma área geográfica, contrato
diferente. A planilha de controle pode trazer as duas famílias de código
pra mesmo lugar: 1/2/3/11/12/13 = manutenção, 14/15/16/22/23/24 =
restauração da MESMA região física (confirmado pela usuária em 2026-08-14).
`MAPA_REGIAO_GEOGRAFICA` traduz o código de restauração pro código
geográfico que a ficha usa — sem isso a O.S. de restauração nunca acharia
geometria (R14 não existe em camadas/, é sempre R1/R2/R3/R11/R12/R13).

Saída (tudo em dados/, consumido pelo index.html sem build):
    dados/os_<REGIAO>.js       -- um por região geográfica (R1, R2, ...)
    relatorio_qualidade_os.txt -- trecho não encontrado no shapefile etc.

Rodar:  python converter_os.py
Requer: openpyxl, geopandas, shapely
"""
import glob
import json
import os
import re
import sys
import unicodedata

import openpyxl
import geopandas as gpd
from shapely.geometry import mapping

BASE = os.path.dirname(os.path.abspath(__file__))
OS_DIR = os.path.join(BASE, 'fichas', 'OS')
CAMADAS_DIR = os.path.join(BASE, 'camadas')  # cópia própria (repo separado, sem depender de web - fichas)
DADOS_DIR = os.path.join(BASE, 'dados')
EPSG_METRICO = 31982  # SIRGAS 2000 / UTM 22S — mesmo do geoportal principal

# restauração (chave) -> manutenção/geográfico (valor) — mesma área física.
MAPA_REGIAO_GEOGRAFICA = {14: 1, 15: 2, 16: 3, 22: 11, 23: 12, 24: 13}
REGIOES_RESTAURACAO = set(MAPA_REGIAO_GEOGRAFICA.keys())

MESES_PT = {1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'}

qa_msgs = []


def qa(msg):
    qa_msgs.append(msg)
    print('  [QA]', msg)


def _norm(s):
    """Maiúsculas, sem acento, sem quebra de linha, espaços colapsados."""
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'\s+', ' ', s).strip().upper()
    return s


def _achar_cabecalho(ws, texto_norm, linhas=range(1, 6)):
    for r in linhas:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and texto_norm in _norm(v):
                return r, c
    return None


# ---------------------------------------------------------------------
# Geometria: camadas/R<região>_TRECHOS.shp, agrupada pelo campo Id.
# Reaproveita o mesmo shapefile que converter_fichas.py usa (uma linha por
# S.R.E.) — aqui não corta por km, pega a linha inteira de cada S.R.E. que
# pertence ao trecho.
# ---------------------------------------------------------------------
_cache_trechos = {}


def carregar_trechos_regiao(regiao_num):
    if regiao_num in _cache_trechos:
        return _cache_trechos[regiao_num]
    caminho = os.path.join(CAMADAS_DIR, f'R{regiao_num}_TRECHOS.shp')
    if not os.path.exists(caminho):
        qa(f'R{regiao_num}: shapefile R{regiao_num}_TRECHOS.shp não encontrado em camadas/')
        _cache_trechos[regiao_num] = {}
        return {}
    gdf = gpd.read_file(caminho)
    if gdf.crs is None:
        gdf = gdf.set_crs(EPSG_METRICO)
    gdf = gdf.to_crs(4326)
    col_id = next((c for c in gdf.columns if _norm(c) == 'ID'), None)
    if col_id is None:
        qa(f'R{regiao_num}: shapefile R{regiao_num}_TRECHOS.shp sem coluna Id reconhecível (colunas: {list(gdf.columns)})')
        _cache_trechos[regiao_num] = {}
        return {}
    por_trecho = {}
    for _, row in gdf.iterrows():
        try:
            n = int(row[col_id])
        except (TypeError, ValueError):
            continue
        if row.geometry is None or row.geometry.is_empty:
            continue
        por_trecho.setdefault(n, []).append(row.geometry)
    _cache_trechos[regiao_num] = por_trecho
    return por_trecho


# ---------------------------------------------------------------------
# Leitura das abas "REGIÃO XX - HISTÓRICO" — medição mensal por O.S.P.
# (VALOR PREVISTO/TOTAL MEDIDO/SALDO/% já vêm calculados pela planilha;
# aqui só lemos, não recalculamos). Chave: número da O.S.P. dentro da
# mesma aba de região (o "N°" da planilha) — histórico não guarda o
# CONTRATO, mas OSP já é único dentro da região/planilha.
# ---------------------------------------------------------------------
def parse_historico_os(wb, regiao_num_planilha):
    nome_aba = f'REGIÃO {regiao_num_planilha:02d} - HISTÓRICO'
    if nome_aba not in wb.sheetnames:
        return {}
    ws = wb[nome_aba]

    # Colunas fixas: B=ID, C=OSP, D=CRONOGRAMA, E=VALOR PREVISTO,
    # F=TOTAL MEDIDO, G=SALDO, H=%, I=SITUAÇÃO FINAL, depois 12 pares
    # (VALOR MEDIDO, SITUAÇÃO) — um por mês, com o mês em datetime na
    # linha 2 (célula mesclada com a coluna de SITUAÇÃO).
    meses_cols = []
    for c in range(10, 34, 2):
        dt = ws.cell(row=2, column=c).value
        meses_cols.append((c, dt.month if dt else None))

    def num(v):
        return v if isinstance(v, (int, float)) else None

    resultado = {}
    for r in range(4, ws.max_row + 1):
        osp = ws.cell(row=r, column=3).value
        if not isinstance(osp, (int, float)) or osp == 0:
            continue
        situacao_final = ws.cell(row=r, column=9).value
        medido_mensal = []
        for col_valor, mes_num in meses_cols:
            if mes_num is None:
                continue
            v = num(ws.cell(row=r, column=col_valor).value) or 0
            medido_mensal.append({'mes': MESES_PT[mes_num], 'valor': v})
        resultado[int(osp)] = {
            'valor_executado': num(ws.cell(row=r, column=6).value) or 0,
            'saldo': num(ws.cell(row=r, column=7).value),
            'pct_executado': num(ws.cell(row=r, column=8).value),  # None se #DIV/0! etc.
            'situacao_final': str(situacao_final).strip() if situacao_final not in (None, '', '-') else None,
            'medido_mensal': medido_mensal,
        }
    return resultado


# ---------------------------------------------------------------------
# Leitura das abas "REGIÃO XX - RESUMO"
# ---------------------------------------------------------------------
def parse_planilha_os(caminho):
    print(f'Lendo {os.path.basename(caminho)} ...')
    wb = openpyxl.load_workbook(caminho, data_only=True, keep_vba=False)
    resultados_por_regiao = {}  # 'R1' -> [feature, feature, ...]
    cache_historico = {}  # regiao_num_planilha -> {osp: {...}}

    for nome_aba in wb.sheetnames:
        if 'RESUMO' not in _norm(nome_aba):
            continue
        ws = wb[nome_aba]

        pos_id = _achar_cabecalho(ws, 'ID')
        pos_situacao = _achar_cabecalho(ws, 'SITUACAO')
        if not (pos_id and pos_situacao):
            print(f'  [aviso] aba "{nome_aba}" não parece uma aba de resumo de O.S.P. — pulando')
            continue
        linha_cab, col_id = pos_id
        cols = {}
        for c in range(1, ws.max_column + 1):
            titulo = _norm(ws.cell(row=linha_cab, column=c).value)
            if titulo:
                cols[titulo] = c

        def valor(row, *chaves):
            for k in chaves:
                if k in cols:
                    return ws.cell(row=row, column=cols[k]).value
            return None

        n_lidos = 0
        n_pulados_sem_trecho = 0
        n_sem_geometria = 0

        for r in range(linha_cab + 1, ws.max_row + 1):
            osp = valor(r, 'OSP')
            regiao_txt = valor(r, 'REGIAO')
            if osp in (None, '', 0) or regiao_txt in (None, ''):
                continue
            if isinstance(osp, str) and 'REF' in osp.upper():
                qa(f'{nome_aba}, linha {r}: registro com erro de fórmula (#REF!) na planilha — pulando')
                continue

            m = re.search(r'(\d+)', str(regiao_txt))
            if not m:
                continue
            regiao_num_planilha = int(m.group(1))
            tipo_servico = 'restauracao' if regiao_num_planilha in REGIOES_RESTAURACAO else 'manutencao'
            regiao_geo = MAPA_REGIAO_GEOGRAFICA.get(regiao_num_planilha, regiao_num_planilha)
            regiao_geo_label = f'R{regiao_geo}'

            trecho_num = valor(r, 'N TRECHO')
            trecho_nome = valor(r, 'TRECHO')
            sem_trecho_cadastrado = not isinstance(trecho_num, (int, float)) or trecho_num in (0,) or (
                isinstance(trecho_nome, str) and 'NAO CADASTRADO' in _norm(trecho_nome)
            )
            if sem_trecho_cadastrado:
                n_pulados_sem_trecho += 1
                trecho_num = None
            else:
                trecho_num = int(trecho_num)

            situacao = valor(r, 'SITUACAO')
            situacao = str(situacao).strip() if situacao not in (None, '', 0) else 'Não informada'

            if regiao_num_planilha not in cache_historico:
                cache_historico[regiao_num_planilha] = parse_historico_os(wb, regiao_num_planilha)
            hist = cache_historico[regiao_num_planilha].get(int(osp)) if isinstance(osp, (int, float)) else None

            # A O.S.P. é um registro administrativo real mesmo quando não dá pra
            # desenhar no mapa (sem trecho cadastrado ainda, ou trecho não bate
            # com o shapefile) — ela CONTINUA entrando na lista/KPIs/gráficos
            # (senão os totais do geoportal nunca batem com o dashboard da
            # planilha), só não ganha geometria. Mesmo critério já usado no
            # geoportal de fichas (ver [[geoportal-fichas-inspecao]]).
            geoms = None
            if not sem_trecho_cadastrado:
                geoms = carregar_trechos_regiao(regiao_geo).get(trecho_num)
                if not geoms:
                    qa(f'{regiao_geo_label} (via {nome_aba}): Trecho {trecho_num} ({trecho_nome}) não encontrado '
                       f'no shapefile R{regiao_geo}_TRECHOS.shp — mantido na lista/KPIs, sem geometria no mapa')
                    n_sem_geometria += 1

            contrato = valor(r, 'CONTRATO')
            # A planilha só guarda o MÊS na "DATA EMISSÃO" ("Janeiro", sem ano) —
            # e o mesmo mês se repete em contratos de anos diferentes (ex.: R03
            # tem "Julho" tanto no contrato 002.2025 quanto no 1452.2026). O
            # número do contrato sempre termina em ".AAAA" — usa isso como ano de
            # emissão (o contrato dura ~1 ano, então o mês só pode ser daquele
            # ano). Vira "Julho/2025" em vez de só "Julho", pra não misturar anos.
            mes_emissao = valor(r, 'DATA EMISSAO')
            m_ano = re.search(r'(\d{4})\s*$', str(contrato or ''))
            if mes_emissao and mes_emissao != '-' and m_ano:
                data_emissao = f'{mes_emissao}/{m_ano.group(1)}'
            else:
                data_emissao = mes_emissao

            props = {
                'regiao': regiao_geo_label,
                'regiao_os': f'R{regiao_num_planilha:02d}',
                'tipo_servico': tipo_servico,
                'contrato': contrato,
                'osp': osp,
                'data_emissao': data_emissao,
                'cronograma': valor(r, 'CRONOGRAMA'),
                'trecho_num': trecho_num,
                'trecho_nome': trecho_nome,
                'servico': valor(r, 'SERVICO'),
                'valor_previsto': valor(r, 'VALOR PREVISTO') or 0,
                'situacao': situacao,
                'observacao': valor(r, 'OBSERVACAO'),
                # Vindos da aba HISTÓRICO (medição mensal) — None/0 se a O.S.P.
                # ainda não tem histórico lançado (ex. recém emitida).
                'valor_executado': hist['valor_executado'] if hist else 0,
                'saldo': hist['saldo'] if hist else None,
                'pct_executado': hist['pct_executado'] if hist else None,
                'situacao_final': hist['situacao_final'] if hist else None,
                'medido_mensal': hist['medido_mensal'] if hist else [],
            }
            if geoms:
                for geom in geoms:
                    resultados_por_regiao.setdefault(regiao_geo_label, []).append({
                        'type': 'Feature',
                        'geometry': mapping(geom),
                        'properties': props,
                    })
            else:
                resultados_por_regiao.setdefault(regiao_geo_label, []).append({
                    'type': 'Feature',
                    'geometry': None,
                    'properties': props,
                })
            n_lidos += 1

        print(f'  {nome_aba}: {n_lidos} O.S.P. lida(s), {n_pulados_sem_trecho} sem trecho cadastrado ainda, '
              f'{n_sem_geometria} sem geometria no shapefile')

    return resultados_por_regiao


def main():
    arquivos = sorted(
        f for f in glob.glob(os.path.join(OS_DIR, '*.xls*'))
        if not os.path.basename(f).startswith('~$')  # lock file do Excel (planilha aberta)
    )
    if not arquivos:
        print(f'Nenhuma planilha de O.S.P. encontrada em {OS_DIR}')
        return

    geral = {}  # 'R1' -> [features]
    for caminho in arquivos:
        try:
            parcial = parse_planilha_os(caminho)
        except Exception as e:
            qa(f'{os.path.basename(caminho)}: falha ao ler ({e}) — arquivo ignorado')
            continue
        for regiao, feats in parcial.items():
            geral.setdefault(regiao, []).extend(feats)

    os.makedirs(DADOS_DIR, exist_ok=True)
    total_osp = 0
    for regiao, feats in sorted(geral.items()):
        pacote = {'type': 'FeatureCollection', 'features': feats}
        nome_arquivo = f'os_{regiao}.js'
        with open(os.path.join(DADOS_DIR, nome_arquivo), 'w', encoding='utf-8') as f:
            f.write('// Gerado por converter_os.py — não editar à mão\n')
            f.write('window.DADOS_OS_REGIAO = window.DADOS_OS_REGIAO || {};\n')
            f.write(f'window.DADOS_OS_REGIAO["{regiao}"] = {json.dumps(pacote, ensure_ascii=False)};\n')
        n_osp = len(set((f['properties']['contrato'], f['properties']['osp'], f['properties']['trecho_num'])
                        for f in feats))
        total_osp += n_osp
        print(f'  -> dados/{nome_arquivo} ({len(feats)} feature(s), {n_osp} O.S.P.+trecho)')

    with open(os.path.join(DADOS_DIR, 'manifest_os.js'), 'w', encoding='utf-8') as f:
        f.write('// Gerado por converter_os.py — não editar à mão\n')
        f.write('window.MANIFEST_OS = ' + json.dumps(sorted(geral.keys()), ensure_ascii=False) + ';\n')

    with open(os.path.join(BASE, 'relatorio_qualidade_os.txt'), 'w', encoding='utf-8') as f:
        if qa_msgs:
            f.write('\n'.join(qa_msgs) + '\n')
        else:
            f.write('Nenhum problema encontrado.\n')

    print(f'\n{total_osp} O.S.P.+trecho no total, {len(qa_msgs)} observação(ões) — ver relatorio_qualidade_os.txt')


if __name__ == '__main__':
    main()
